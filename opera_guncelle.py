#!/usr/bin/env python3
"""
Opera Analiz — EPİAŞ Şeffaflık verilerini Excel'e işler.
"Opera Enerji Veri" sekmesinde ayın başından bugüne (yarının PTF'si açıklandıysa yarına):
  F: PTF (TL/MWh)   G: SMF   H: PTF (USD/MWh)   I: AOF (GİP AOF)
  Santral blokları (UEVM ← gerçek zamanlı üretim, İLK KGÜP ← kgup-v1, SON KGÜP ← kgup):
    ARPACIK HES  → AB / AC / AF     YAVUZ HES → AI / AJ / AM     MİDİLLİ HES → AP / AQ / AT
Ayrıca Tarih (A) ve Saat (B) hücreleri gerçek değerle yazılır (kopuk dış-bağlantı formülleri yerine).

Ayarlar (API kullanıcısı/şifresi, Excel klasörü, dosya adı kalıbı): ayarlar.json —
yoksa ayarlar.ornek.json'u o adla kopyalayıp doldur. Dosya adı içinde bulunulan aya göre
otomatik türetilir: "Opera Analiz {AY} {YIL}.xlsx" → "Opera Analiz Agustos 2026.xlsx"
(Türkçe ay adı, İngilizce karakterlerle; gerçek Türkçe yazım da denenir).

Kullanım:
  python opera_guncelle.py                 # sadece Excel'i güncelle
  python opera_guncelle.py --push          # + kopyasını repoya koy, commit'le, GitHub'a pushla
  python opera_guncelle.py --ay 2026-08    # başka ay (varsayılan: içinde bulunulan ay)
  python opera_guncelle.py --dosya "C:\\...\\baska.xlsx"   # kalıbı ezmek için

Şifre repoya asla girmez (ayarlar.json gitignore'da).
NOT: Excel dosyası açıkken kaydedilemez — Excel'i kapatıp yeniden çalıştır.
"""
import argparse, datetime as dt, json, os, shutil, subprocess, sys
from pathlib import Path
import pandas as pd
import openpyxl
from eptr2 import EPTR2

HERE = Path(__file__).resolve().parent
SEKME = "Opera Enerji Veri"
ILK_SATIR = 3                      # veri 3. satırdan başlar (başlık 2. satır)

AY_ADI = {1: "Ocak", 2: "Subat", 3: "Mart", 4: "Nisan", 5: "Mayis", 6: "Haziran",
          7: "Temmuz", 8: "Agustos", 9: "Eylul", 10: "Ekim", 11: "Kasim", 12: "Aralik"}
AY_ADI_TR = {2: "Şubat", 5: "Mayıs", 8: "Ağustos", 9: "Eylül", 11: "Kasım", 12: "Aralık"}

def ayarlari_yukle():
    for ad in ("ayarlar.json", "ayarlar.ornek.json"):
        p = HERE / ad
        if p.exists():
            return {k: v for k, v in json.loads(p.read_text(encoding="utf-8")).items()
                    if not k.startswith("_")}
    sys.exit("HATA: ayarlar.json yok (ayarlar.ornek.json'u kopyalayıp doldurun).")

def yol_cevir(s):
    """Windows yolu (C:\\...) → WSL yolu (/mnt/c/...)."""
    s = s.strip().replace("\\", "/")
    if len(s) > 1 and s[1] == ":":
        s = f"/mnt/{s[0].lower()}{s[2:]}"
    return Path(s)

def sifre_bul(ayar):
    if os.environ.get("EPIAS_PW"): return os.environ["EPIAS_PW"]
    if ayar.get("api_sifre"): return ayar["api_sifre"]
    f = Path(ayar.get("api_sifre_dosyasi", "~/.epias_pw")).expanduser()
    if f.exists(): return f.read_text(encoding="utf-8").splitlines()[0].strip()
    sys.exit("HATA: şifre yok — EPIAS_PW değişkeni, ayarlar.json api_sifre ya da şifre dosyası gerekli.")

def dosya_bul(ayar, yil, ay):
    klasor = yol_cevir(ayar["dosya_klasoru"])
    kalip = ayar.get("dosya_adi_kalibi", "Opera Analiz {AY} {YIL}.xlsx")
    adaylar = [kalip.replace("{AY}", AY_ADI[ay]).replace("{YIL}", str(yil))]
    if ay in AY_ADI_TR:  # gerçek Türkçe yazımla da dene (Ağustos, Eylül...)
        adaylar.append(kalip.replace("{AY}", AY_ADI_TR[ay]).replace("{YIL}", str(yil)))
    for ad in adaylar:
        if (klasor / ad).exists(): return klasor / ad
    sys.exit(f"HATA: dosya bulunamadı: {klasor} içinde {' veya '.join(adaylar)}")

ORG_ID = "104782"                  # OPERA ENERJİ A.Ş. (TOPLAYICI) — 40X000000104782K
# ad → (rt-gen pp_id, kgup uevcb_id, [UEVM, İLK KGÜP, SON KGÜP] sütunları)
SANTRALLER = {
    "ARPACIK HES": ("1942", "3194132", ["AB", "AC", "AF"]),
    "YAVUZ HES":   ("1263", "4138",    ["AI", "AJ", "AM"]),
    "MİDİLLİ HES": ("1265", "117864",  ["AP", "AQ", "AT"]),
}

def ts_dict(df, kolon):
    """DataFrame → {(gün, saat): değer} — date kolonu ISO+03:00 gelir."""
    if kolon not in df.columns: return {}
    t = pd.to_datetime(df["date"], utc=True).dt.tz_convert("Etc/GMT-3")
    return {(g.day, g.hour): v for g, v in zip(t, df[kolon]) if pd.notna(v)}

def cek(e, key, bas, bit, **kw):
    """Aralığı çek; uç tarih reddedilirse günü geri çekerek dene.
    (santral bazlı rt-gen yalnız DÜNE kadar verir; mcp yarını 14:00'ten sonra verir)"""
    for geri in range(4):
        son = bit - dt.timedelta(days=geri)
        if son < bas: break
        try:
            df = e.call(key, start_date=bas.isoformat(), end_date=son.isoformat(), **kw)
            if len(df): return df
        except Exception:
            continue
    print(f"   uyarı: {key} verisi alınamadı ({bas} → {bit})")
    return pd.DataFrame(columns=["date"])

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ay", default=dt.date.today().strftime("%Y-%m"))
    ap.add_argument("--dosya", default="", help="ayarlar.json kalıbını ezmek için tam yol")
    ap.add_argument("--push", action="store_true", help="kopyayı repoya koy + git push")
    ap.add_argument("--sifirdan", action="store_true",
                    help="tüm veri satırlarını komple temizleyip baştan doldur")
    a = ap.parse_args()

    ayar = ayarlari_yukle()
    yil, ay = map(int, a.ay.split("-"))
    bas = dt.date(yil, ay, 1)
    bugun = dt.date.today()
    bit = min(bugun + dt.timedelta(days=1),                       # yarının PTF'si 14:00'ten sonra var
              (bas + dt.timedelta(days=40)).replace(day=1) - dt.timedelta(days=1))
    if bit < bas: sys.exit(f"HATA: {a.ay} henüz başlamadı.")
    if a.dosya:
        xlsx = yol_cevir(a.dosya)
        if not xlsx.exists(): sys.exit(f"HATA: dosya yok: {xlsx}")
    else:
        xlsx = dosya_bul(ayar, yil, ay)

    e = EPTR2(username=ayar["api_kullanici"], password=sifre_bul(ayar))
    print(f"Aralık: {bas} → {bit}  |  Dosya: {xlsx.name}")

    print("-> fiyatlar (mcp/smp/wap) + sistem yönü (mcp-smp-imb)")
    mcp = cek(e, "mcp", bas, bit); smp = cek(e, "smp", bas, bit); wap = cek(e, "wap", bas, bit)
    imb = cek(e, "mcp-smp-imb", bas, bit)
    print("-> Opera org: GÖP eşleşme (dam-clearing) + İA miktarları (bi-short/bi-long)")
    gop = cek(e, "dam-clearing", bas, bit, org_id=ORG_ID)
    ia_sat = cek(e, "bi-short", bas, bit, org_id=ORG_ID)   # İA satış miktarı
    ia_al  = cek(e, "bi-long",  bas, bit, org_id=ORG_ID)   # İA alış miktarı
    seriler = {"E": ts_dict(imb, "systemStatus"),
               "F": ts_dict(mcp, "price"), "G": ts_dict(smp, "systemMarginalPrice"),
               "H": ts_dict(mcp, "priceUsd"), "I": ts_dict(wap, "wap"),
               "J": ts_dict(gop, "matchedOffers"),          # GÖP SSM (satış eşleşme)
               "L": ts_dict(gop, "matchedBids"),            # GÖP SAM (alış eşleşme)
               "N": ts_dict(ia_sat, "quantity"), "Q": ts_dict(ia_al, "quantity")}
    for ad, (pp_id, uevcb_id, kolonlar) in SANTRALLER.items():
        print(f"-> {ad} (uevm + ilk/son kgüp)")
        seriler[kolonlar[0]] = ts_dict(cek(e, "rt-gen", bas, bit, pp_id=pp_id), "total")
        seriler[kolonlar[1]] = ts_dict(cek(e, "kgup-v1", bas, bit, org_id=ORG_ID, uevcb_id=uevcb_id), "toplam")
        seriler[kolonlar[2]] = ts_dict(cek(e, "kgup", bas, bit, org_id=ORG_ID, uevcb_id=uevcb_id), "toplam")

    from openpyxl.formula.translate import Translator
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[SEKME]
    gun_sayisi = ((bas + dt.timedelta(days=40)).replace(day=1) - bas).days
    son_satir = ILK_SATIR + gun_sayisi * 24 - 1

    # kopuk dış-bağlantı formüllerini temizle: '[1]...' başka bilgisayardaki kaynak dosyaya
    # bakıyordu, bir daha hesaplanamaz — Excel'de 0/#REF çöpü olarak görünüyordu.
    temiz = 0
    for r in range(ILK_SATIR, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("=") and "[1]" in v:
                ws.cell(r, c).value = None; temiz += 1
    if temiz: print(f"Temizlenen kopuk formül hücresi: {temiz}")

    # satır iskeleti (K=F*J, O==F gibi şirket-içi formüller): yeni gün dolarken boş hücreler
    # DÜNÜN AYNI SAATİNDEN kopyalanır (formüller satıra çevrilir); dün de boşsa satır 3
    # şablonundan. Şablon, --sifirdan temizliğinden ÖNCE alınır ki iskelet kaybolmasın.
    yonetilen_c = [openpyxl.utils.column_index_from_string(k) for k in ({"A", "B"} | set(seriler))]
    sablon = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(ILK_SATIR, c).value
        if c not in yonetilen_c and v is not None: sablon[c] = v

    if a.sifirdan:
        n = sum(1 for r in range(ILK_SATIR, ws.max_row + 1) for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None)
        for r in range(ILK_SATIR, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value is not None: ws.cell(r, c).value = None
        print(f"Tam temizlik: {n} hücre sıfırlandı, baştan dolduruluyor.")

    def iskelet(r):
        kaynak = r - 24
        for c in range(1, ws.max_column + 1):
            if c in yonetilen_c or ws.cell(r, c).value is not None: continue
            v, org_r = (ws.cell(kaynak, c).value, kaynak) if kaynak >= ILK_SATIR else (None, None)
            if v is None: v, org_r = sablon.get(c), ILK_SATIR
            if v is None: continue
            kol = openpyxl.utils.get_column_letter(c)
            ws.cell(r, c).value = (Translator(v, origin=f"{kol}{org_r}").translate_formula(f"{kol}{r}")
                                   if isinstance(v, str) and v.startswith("=") else v)

    yazilan = {k: 0 for k in seriler}
    for gun in range(1, gun_sayisi + 1):
        for saat in range(24):
            r = ILK_SATIR + (gun - 1) * 24 + saat
            if any((gun, saat) in s for s in seriler.values()):
                ws.cell(r, 1).value = dt.datetime(yil, ay, gun, saat)
                ws.cell(r, 2).value = f"{saat:02d}:00"
                for kolon, s in seriler.items():
                    if (gun, saat) in s:
                        v = s[(gun, saat)]
                        ws[f"{kolon}{r}"].value = v if isinstance(v, str) else float(v)
                        yazilan[kolon] += 1
                iskelet(r)                   # boş kalan iskelet hücrelerini dünden tamamla
            else:                            # verisi olmayan satır KOMPLE boş kalsın
                for c in range(1, ws.max_column + 1):
                    if ws.cell(r, c).value is not None: ws.cell(r, c).value = None
    for r in range(son_satir + 1, ws.max_row + 1):   # ay bloğu sonrası artık satırlar
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None: ws.cell(r, c).value = None
    try:
        wb.save(xlsx)
    except PermissionError:
        sys.exit("HATA: Excel dosyası AÇIK — kapatıp yeniden çalıştırın (kaydedilemedi).")
    ozet = ", ".join(f"{k}:{n}" for k, n in yazilan.items())
    print(f"Yazıldı ({xlsx.name}) — hücre sayıları: {ozet}")

    if a.push:
        kopya = HERE / xlsx.name
        if xlsx.resolve() != kopya.resolve() and \
           (not kopya.exists() or xlsx.stat().st_mtime > kopya.stat().st_mtime):
            shutil.copy2(xlsx, kopya)   # yalnız daha yeniyse kopyala (bayat, tazeyi ezmesin)
        for cmd in (["git", "add", "-A"],
                    ["git", "commit", "-m", f"Veri güncellemesi {dt.datetime.now():%Y-%m-%d %H:%M}"],
                    ["git", "push"]):
            r = subprocess.run(cmd, cwd=HERE, capture_output=True, text=True)
            if r.returncode != 0 and "nothing to commit" not in (r.stdout + r.stderr):
                print(f"git uyarı ({' '.join(cmd)}): {(r.stderr or r.stdout).strip()[:200]}")
        print("GitHub'a pushlandı.")

if __name__ == "__main__":
    main()
